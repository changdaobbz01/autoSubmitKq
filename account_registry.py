from __future__ import annotations

import json
import re
import time
from copy import deepcopy
from dataclasses import dataclass
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from attendance_auth_client import DEFAULT_SESSION_PATH, AttendanceAuthClient, SessionStore
from runtime_paths import APP_ROOT

PROJECT_DIR = APP_ROOT
AUTH_DIR = PROJECT_DIR / ".attendance_auth"
ACCOUNTS_DIR = AUTH_DIR / "accounts"
REGISTRY_PATH = AUTH_DIR / "accounts_registry.json"


HEADER_ALIASES = {
    "userAccount": {
        "useraccount",
        "account",
        "username",
        "login",
        "\u8d26\u53f7",
        "\u5e10\u53f7",
        "\u7528\u6237\u540d",
    },
    "password": {
        "password",
        "pwd",
        "pass",
        "\u5bc6\u7801",
    },
    "realName": {
        "realname",
        "name",
        "\u59d3\u540d",
        "\u771f\u5b9e\u59d3\u540d",
    },
    "department": {
        "department",
        "dept",
        "\u90e8\u95e8",
    },
    "enabled": {
        "enabled",
        "enable",
        "active",
        "\u542f\u7528",
        "\u662f\u5426\u542f\u7528",
    },
    "note": {
        "note",
        "remark",
        "\u5907\u6ce8",
        "\u8bf4\u660e",
    },
    "photoPath": {
        "photopath",
        "photo",
        "imagepath",
        "image",
        "avatar",
        "\u7167\u7247",
        "\u7167\u7247\u5730\u5740",
        "\u56fe\u7247",
        "\u56fe\u7247\u5730\u5740",
        "\u6253\u5361\u7167\u7247",
    },
}


def normalize_header(value: Any) -> str:
    return str(value or "").strip().replace(" ", "").replace("_", "").lower()


def normalize_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = str(value or "").strip().lower()
    if text in {"", "1", "true", "yes", "y", "\u542f\u7528", "\u662f"}:
        return True
    if text in {"0", "false", "no", "n", "\u505c\u7528", "\u5426"}:
        return False
    return True


def mask_password(password: str) -> str:
    if not password:
        return ""
    if len(password) <= 2:
        return "*" * len(password)
    return f"{password[0]}{'*' * (len(password) - 2)}{password[-1]}"


def safe_account_dir_name(user_account: str) -> str:
    return re.sub(r"[^a-zA-Z0-9._-]+", "_", user_account).strip("._-") or "account"


def format_epoch(epoch: int | None) -> str:
    if not epoch:
        return ""
    return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(epoch))


def build_photo_state(photo_path: str) -> dict[str, Any]:
    raw_path = str(photo_path or "").strip()
    if not raw_path:
        return {
            "configured": False,
            "exists": False,
            "status": "missing",
            "statusText": "\u672a\u914d\u7f6e\u7167\u7247",
            "path": "",
            "size": None,
        }

    candidate = Path(raw_path).expanduser()
    try:
        exists = candidate.exists() and candidate.is_file()
    except OSError:
        return {
            "configured": True,
            "exists": False,
            "status": "invalid",
            "statusText": "\u7167\u7247\u8def\u5f84\u4e0d\u53ef\u8bbf\u95ee",
            "path": raw_path,
            "size": None,
        }

    if not exists:
        return {
            "configured": True,
            "exists": False,
            "status": "not_found",
            "statusText": "\u7167\u7247\u8def\u5f84\u4e0d\u5b58\u5728",
            "path": raw_path,
            "size": None,
        }

    return {
        "configured": True,
        "exists": True,
        "status": "ready",
        "statusText": "\u7167\u7247\u53ef\u7528",
        "path": str(candidate),
        "size": candidate.stat().st_size,
    }


@dataclass
class ImportedAccount:
    user_account: str
    password: str
    real_name: str
    department: str
    enabled: bool
    note: str
    photo_path: str
    row_number: int


class AccountRegistry:
    def __init__(self, registry_path: Path = REGISTRY_PATH, accounts_dir: Path = ACCOUNTS_DIR) -> None:
        self.registry_path = registry_path
        self.accounts_dir = accounts_dir

    def load(self) -> dict[str, Any]:
        if not self.registry_path.exists():
            return {"accounts": []}
        try:
            payload = json.loads(self.registry_path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {"accounts": []}
        if not isinstance(payload, dict):
            return {"accounts": []}
        accounts = payload.get("accounts")
        if not isinstance(accounts, list):
            payload["accounts"] = []
        return payload

    def save(self, payload: dict[str, Any]) -> None:
        self.registry_path.parent.mkdir(parents=True, exist_ok=True)
        temp_path = self.registry_path.with_suffix(".tmp")
        temp_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        temp_path.replace(self.registry_path)

    def list_accounts(self, include_sensitive: bool = False) -> list[dict[str, Any]]:
        payload = self.load()
        records: list[dict[str, Any]] = []
        for account in payload.get("accounts", []):
            if not isinstance(account, dict):
                continue
            records.append(self._build_account_summary(account, include_sensitive=include_sensitive))
        return records

    def get_enabled_accounts(self, include_sensitive: bool = False) -> list[dict[str, Any]]:
        return [item for item in self.list_accounts(include_sensitive=include_sensitive) if item.get("enabled")]

    def import_xlsx(self, file_bytes: bytes) -> dict[str, Any]:
        workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValueError("\u8868\u683c\u4e3a\u7a7a")

        header_row = rows[0]
        header_values = [str(value or "").strip() for value in header_row]
        header_map = self._build_header_map(header_row)
        if "userAccount" not in header_map:
            raise ValueError(
                "\u672a\u627e\u5230\u8d26\u53f7\u5217\uff0c\u81f3\u5c11\u9700\u8981 userAccount / \u8d26\u53f7 / \u7528\u6237\u540d \u4e4b\u4e00"
            )

        imported_rows: list[ImportedAccount] = []
        row_numbers_by_account: dict[str, list[int]] = {}

        for row_number, row in enumerate(rows[1:], start=2):
            if row is None:
                continue
            user_account = str(row[header_map["userAccount"]] or "").strip()
            if not user_account:
                continue
            row_numbers_by_account.setdefault(user_account, []).append(row_number)
            imported_rows.append(
                ImportedAccount(
                    user_account=user_account,
                    password=self._read_cell(row, header_map, "password"),
                    real_name=self._read_cell(row, header_map, "realName"),
                    department=self._read_cell(row, header_map, "department"),
                    enabled=normalize_bool(row[header_map["enabled"]]) if "enabled" in header_map else True,
                    note=self._read_cell(row, header_map, "note"),
                    photo_path=self._read_cell(row, header_map, "photoPath"),
                    row_number=row_number,
                )
            )

        latest_import_by_account: dict[str, ImportedAccount] = {}
        for imported in imported_rows:
            latest_import_by_account[imported.user_account] = imported

        duplicate_accounts_in_file = [
            {"userAccount": user_account, "rows": row_numbers}
            for user_account, row_numbers in sorted(row_numbers_by_account.items())
            if len(row_numbers) > 1
        ]

        registry = self.load()
        by_account = {
            str(item.get("userAccount") or "").strip(): deepcopy(item)
            for item in registry.get("accounts", [])
            if isinstance(item, dict) and str(item.get("userAccount") or "").strip()
        }

        now = int(time.time())
        merged_accounts: list[str] = []
        created_accounts: list[str] = []
        merged_change_details: list[dict[str, Any]] = []

        for imported in latest_import_by_account.values():
            existing = by_account.get(imported.user_account, {})
            if existing:
                merged_accounts.append(imported.user_account)
            else:
                created_accounts.append(imported.user_account)

            old_password = str(existing.get("password") or "")
            old_real_name = str(existing.get("realName") or "")
            old_department = str(existing.get("department") or "")
            old_enabled = bool(existing.get("enabled", True)) if existing else True
            old_note = str(existing.get("note") or "")
            old_photo_path = str(existing.get("photoPath") or "")

            next_password = imported.password or old_password
            next_real_name = imported.real_name or old_real_name
            next_department = imported.department or old_department
            next_enabled = imported.enabled
            next_note = imported.note
            next_photo_path = imported.photo_path or old_photo_path

            updated = {
                **existing,
                "userAccount": imported.user_account,
                "password": next_password,
                "realName": next_real_name,
                "department": next_department,
                "enabled": next_enabled,
                "note": next_note,
                "photoPath": next_photo_path,
                "updatedAt": now,
            }
            if "createdAt" not in updated:
                updated["createdAt"] = now
            by_account[imported.user_account] = updated
            self._sync_default_session_if_same_account(imported.user_account)

            if existing:
                changed_fields: list[str] = []
                if old_password != next_password:
                    changed_fields.append("password")
                if old_photo_path != next_photo_path:
                    changed_fields.append("photoPath")
                if old_real_name != next_real_name:
                    changed_fields.append("realName")
                if old_department != next_department:
                    changed_fields.append("department")
                if old_enabled != next_enabled:
                    changed_fields.append("enabled")
                if old_note != next_note:
                    changed_fields.append("note")
                if changed_fields:
                    merged_change_details.append(
                        {
                            "userAccount": imported.user_account,
                            "changedFields": changed_fields,
                        }
                    )

        registry["accounts"] = sorted(by_account.values(), key=lambda item: str(item.get("userAccount") or ""))
        registry["updatedAt"] = now
        self.save(registry)

        imported_account_summaries = [
            self.get_account(user_account, include_sensitive=False)
            for user_account in sorted(latest_import_by_account.keys())
        ]
        photo_issues = [
            {
                "userAccount": item["userAccount"],
                "status": item["photo"]["status"],
                "statusText": item["photo"]["statusText"],
                "path": item["photo"]["path"],
            }
            for item in imported_account_summaries
            if not item.get("photo", {}).get("exists")
        ]

        warnings: list[str] = []
        if duplicate_accounts_in_file:
            warnings.append(
                f"\u8868\u683c\u5185\u53d1\u73b0 {len(duplicate_accounts_in_file)} \u4e2a\u91cd\u590d\u8d26\u53f7\uff0c\u5df2\u6309\u6700\u540e\u4e00\u884c\u8986\u76d6\u3002"
            )
        if merged_accounts:
            warnings.append(
                f"\u53d1\u73b0 {len(merged_accounts)} \u4e2a\u5df2\u5b58\u5728\u8d26\u53f7\uff0c\u5df2\u6267\u884c\u5408\u5e76\u66f4\u65b0\u3002"
            )
        if photo_issues:
            warnings.append(
                f"\u68c0\u6d4b\u5230 {len(photo_issues)} \u4e2a\u8d26\u53f7\u7684\u7167\u7247\u8def\u5f84\u4e0d\u53ef\u7528\uff0c\u771f\u5b9e\u6253\u5361\u548c\u5e26\u56fe\u8c03\u6d4b\u4f1a\u5931\u8d25\u6216\u8df3\u8fc7\u3002"
            )

        ready_photo_count = sum(1 for item in imported_account_summaries if item.get("photo", {}).get("exists"))
        password_changed_count = sum(1 for item in merged_change_details if "password" in item["changedFields"])
        photo_path_changed_count = sum(1 for item in merged_change_details if "photoPath" in item["changedFields"])
        if password_changed_count:
            warnings.append(f"\u540c\u8d26\u53f7\u5bc6\u7801\u5df2\u66f4\u65b0 {password_changed_count} \u4e2a\u3002")
        if photo_path_changed_count:
            warnings.append(f"\u540c\u8d26\u53f7\u7167\u7247\u8def\u5f84\u5df2\u66f4\u65b0 {photo_path_changed_count} \u4e2a\u3002")

        return {
            "importedCount": len(latest_import_by_account),
            "sourceRowCount": len(imported_rows),
            "createdCount": len(created_accounts),
            "mergedCount": len(merged_accounts),
            "mergedChangeDetails": merged_change_details,
            "passwordChangedCount": password_changed_count,
            "photoPathChangedCount": photo_path_changed_count,
            "duplicateAccountCount": len(duplicate_accounts_in_file),
            "duplicateAccountsInFile": duplicate_accounts_in_file,
            "mergedAccounts": merged_accounts,
            "createdAccounts": created_accounts,
            "warnings": warnings,
            "photoReadyCount": ready_photo_count,
            "photoProblemCount": len(photo_issues),
            "photoIssues": photo_issues,
            "sheetInfo": {
                "activeSheet": sheet.title,
                "sheetNames": workbook.sheetnames,
                "headerRow": header_values,
                "detectedColumns": sorted(header_map.keys()),
                "dataRowCount": len(imported_rows),
                "uniqueAccountCount": len(latest_import_by_account),
            },
            "accounts": self.list_accounts(),
        }

    def set_enabled(self, user_account: str, enabled: bool) -> dict[str, Any]:
        registry = self.load()
        updated = False
        for account in registry.get("accounts", []):
            if str(account.get("userAccount") or "").strip() == user_account:
                account["enabled"] = enabled
                account["updatedAt"] = int(time.time())
                updated = True
                break
        if not updated:
            raise ValueError(f"\u8d26\u53f7\u4e0d\u5b58\u5728: {user_account}")
        self.save(registry)
        return self._find_account_summary(user_account)

    def set_password(self, user_account: str, password: str) -> dict[str, Any]:
        registry = self.load()
        updated = False
        for account in registry.get("accounts", []):
            if str(account.get("userAccount") or "").strip() == user_account:
                account["password"] = password
                account["updatedAt"] = int(time.time())
                updated = True
                break
        if not updated:
            raise ValueError(f"\u8d26\u53f7\u4e0d\u5b58\u5728: {user_account}")
        self.save(registry)
        return self._find_account_summary(user_account)

    def remove(self, user_account: str) -> dict[str, Any]:
        registry = self.load()
        registry["accounts"] = [
            item
            for item in registry.get("accounts", [])
            if str(item.get("userAccount") or "").strip() != user_account
        ]
        registry["updatedAt"] = int(time.time())
        self.save(registry)
        return {"removed": True, "userAccount": user_account, "accounts": self.list_accounts()}

    def clear_all_tokens(self) -> dict[str, Any]:
        registry = self.load()
        cleared_accounts: list[str] = []
        cleared_count = 0

        for account in registry.get("accounts", []):
            user_account = str(account.get("userAccount") or "").strip()
            if not user_account:
                continue
            session_store = SessionStore(self.get_session_path(user_account))
            if session_store.load() is not None:
                cleared_count += 1
            session_store.clear()
            cleared_accounts.append(user_account)

        default_cleared = False
        default_store = SessionStore(DEFAULT_SESSION_PATH)
        if default_store.load() is not None:
            default_cleared = True
        default_store.clear()

        return {
            "cleared": True,
            "clearedCount": cleared_count,
            "defaultSessionCleared": default_cleared,
            "accounts": cleared_accounts,
            "registry": self.summarize_registry(),
        }

    def update_last_run(self, user_account: str, run_record: dict[str, Any]) -> None:
        registry = self.load()
        for account in registry.get("accounts", []):
            if str(account.get("userAccount") or "").strip() == user_account:
                account["lastRun"] = run_record
                account["updatedAt"] = int(time.time())
                break
        self.save(registry)

    def build_auth_client(self, account: dict[str, Any]) -> AttendanceAuthClient:
        return AttendanceAuthClient(session_store=SessionStore(self.get_session_path(account["userAccount"])))

    def get_session_path(self, user_account: str) -> Path:
        return self.accounts_dir / safe_account_dir_name(user_account) / "session.json"

    def summarize_registry(self) -> dict[str, Any]:
        accounts = self.list_accounts()
        enabled_accounts = [item for item in accounts if item.get("enabled")]
        reusable_accounts = [item for item in enabled_accounts if item.get("session", {}).get("reusable")]
        need_login_accounts = [item for item in enabled_accounts if item.get("needsCaptchaLogin")]
        expired_accounts = [item for item in enabled_accounts if item.get("session", {}).get("status") == "expired"]
        missing_accounts = [item for item in enabled_accounts if item.get("session", {}).get("status") == "missing"]
        ready_photo_accounts = [item for item in enabled_accounts if item.get("photo", {}).get("exists")]
        missing_photo_accounts = [item for item in enabled_accounts if not item.get("photo", {}).get("exists")]
        return {
            "totalCount": len(accounts),
            "enabledCount": len(enabled_accounts),
            "reusableCount": len(reusable_accounts),
            "needsCaptchaLoginCount": len(need_login_accounts),
            "expiredCount": len(expired_accounts),
            "missingTokenCount": len(missing_accounts),
            "photoReadyCount": len(ready_photo_accounts),
            "photoProblemCount": len(missing_photo_accounts),
            "accounts": accounts,
        }

    def get_account(self, user_account: str, include_sensitive: bool = False) -> dict[str, Any]:
        for item in self.list_accounts(include_sensitive=include_sensitive):
            if item.get("userAccount") == user_account:
                return item
        raise ValueError(f"\u8d26\u53f7\u4e0d\u5b58\u5728: {user_account}")

    def _find_account_summary(self, user_account: str) -> dict[str, Any]:
        return self.get_account(user_account, include_sensitive=False)

    def _build_account_summary(self, account: dict[str, Any], include_sensitive: bool = False) -> dict[str, Any]:
        user_account = str(account.get("userAccount") or "").strip()
        password = str(account.get("password") or "")
        photo_path = str(account.get("photoPath") or "").strip()

        session_store = SessionStore(self.get_session_path(user_account))
        client = AttendanceAuthClient(session_store=session_store)
        raw_session = session_store.load()

        if raw_session is None:
            session_payload = {
                "cached": False,
                "reusable": False,
                "status": "missing",
                "statusText": "\u65e0 token",
                "tokenExp": None,
                "tokenExpText": "",
                "savedAtText": "",
            }
        elif client.is_session_reusable(raw_session):
            session_payload = {
                "cached": True,
                "reusable": True,
                "status": "reusable",
                "statusText": "\u53ef\u590d\u7528",
                "tokenExp": raw_session.token_exp,
                "tokenExpText": format_epoch(raw_session.token_exp),
                "savedAtText": format_epoch(raw_session.saved_at),
            }
        else:
            session_payload = {
                "cached": True,
                "reusable": False,
                "status": "expired",
                "statusText": "token \u5df2\u8fc7\u671f",
                "tokenExp": raw_session.token_exp,
                "tokenExpText": format_epoch(raw_session.token_exp),
                "savedAtText": format_epoch(raw_session.saved_at),
            }

        photo_state = build_photo_state(photo_path)
        display_name = str(account.get("realName") or (raw_session.user_name if raw_session else "") or "")
        display_department = str(account.get("department") or (raw_session.department if raw_session else "") or "")
        result = {
            "userAccount": user_account,
            "realName": display_name,
            "department": display_department,
            "enabled": bool(account.get("enabled", True)),
            "note": str(account.get("note") or ""),
            "hasPassword": bool(password),
            "passwordMasked": mask_password(password),
            "needsCaptchaLogin": not session_payload["reusable"],
            "needsCaptchaLoginReason": session_payload["statusText"],
            "session": session_payload,
            "photo": photo_state,
            "photoPath": photo_state["path"] or photo_path,
            "lastRun": account.get("lastRun") if isinstance(account.get("lastRun"), dict) else None,
            "createdAt": account.get("createdAt"),
            "updatedAt": account.get("updatedAt"),
        }
        if include_sensitive:
            result["password"] = password
        return result

    def _build_header_map(self, header_row: tuple[Any, ...]) -> dict[str, int]:
        result: dict[str, int] = {}
        for index, value in enumerate(header_row):
            normalized = normalize_header(value)
            if not normalized:
                continue
            for canonical, aliases in HEADER_ALIASES.items():
                if normalized in aliases and canonical not in result:
                    result[canonical] = index
        return result

    def _read_cell(self, row: tuple[Any, ...], header_map: dict[str, int], key: str) -> str:
        if key not in header_map:
            return ""
        value = row[header_map[key]]
        return str(value or "").strip()

    def _sync_default_session_if_same_account(self, user_account: str) -> None:
        if not DEFAULT_SESSION_PATH.exists():
            return
        default_store = SessionStore(DEFAULT_SESSION_PATH)
        default_session = default_store.load()
        if default_session is None or default_session.user_account != user_account:
            return
        account_store = SessionStore(self.get_session_path(user_account))
        if account_store.load() is None:
            account_store.save(default_session)
